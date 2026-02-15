"""Export functionality for scan results."""

import csv
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .resource_info import get_resource_info, get_safety_display


def export_to_csv(
    resources: list[dict[str, Any]],
    output_dir: Path | str,
    subscription_name: str = "",
    tenant_id: str = "",
) -> Path:
    """Auto-export scan results to CSV file.

    Args:
        resources: List of resource dictionaries
        output_dir: Directory to save the CSV file
        subscription_name: Name of the subscription scanned
        tenant_id: Azure tenant ID for portal URLs

    Returns:
        Path to created CSV file
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Create filename with datetime and subscription
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    safe_sub_name = "".join(c if c.isalnum() or c in "- " else "_" for c in subscription_name)
    safe_sub_name = safe_sub_name.replace(" ", "-")[:50]  # Limit length
    filename = f"scan_{timestamp}_{safe_sub_name}.csv"
    filepath = output_dir / filename

    # Define columns
    columns = ["Name", "Type", "Risk Level", "Safe to Delete", "Resource Group", "Location", "Cost/Month", "Details", "Resource ID", "Portal URL"]

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(columns)

        for resource in resources:
            portal_url = ""
            resource_id = resource.get("id", "")
            if tenant_id and resource_id:
                portal_url = f"https://portal.azure.com/#@{tenant_id}/resource{resource_id}"

            # Get risk level and safety info
            risk_level = resource.get("risk_level", "medium")
            safety_display, _ = get_safety_display(risk_level)
            info = get_resource_info(resource.get("type", ""))
            safe_to_delete = info.get("safe_to_delete", "Verify before deleting.")

            writer.writerow([
                resource.get("name", ""),
                resource.get("type_display", ""),
                risk_level.upper(),
                safe_to_delete,
                resource.get("resource_group", ""),
                resource.get("location", ""),
                resource.get("cost_display", "$0"),
                resource.get("details", ""),
                resource_id,
                portal_url,
            ])

    return filepath


def export_to_excel(
    resources: list[dict[str, Any]],
    filepath: Path | str | None = None,
    tenant_id: str = "",
) -> Path:
    """Export resources to Excel file.

    Args:
        resources: List of resource dictionaries
        filepath: Output file path (default: azure-orphans-YYYY-MM-DD.xlsx)
        tenant_id: Azure tenant ID for portal URLs

    Returns:
        Path to created Excel file
    """
    if filepath is None:
        date_str = datetime.now().strftime("%Y-%m-%d")
        filepath = Path(f"azure-orphans-{date_str}.xlsx")
    else:
        filepath = Path(filepath)

    wb = Workbook()
    ws = wb.active
    ws.title = "Orphaned Resources"

    # Define columns
    columns = ["Name", "Type", "Risk Level", "Safe to Delete", "Resource Group", "Location", "Cost/Month", "Details", "Portal URL"]

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")

    # Risk level fills
    risk_fills = {
        "low": PatternFill(start_color="4A7C4E", end_color="4A7C4E", fill_type="solid"),
        "medium": PatternFill(start_color="B8860B", end_color="B8860B", fill_type="solid"),
        "high": PatternFill(start_color="C75050", end_color="C75050", fill_type="solid"),
    }
    risk_font = Font(bold=True, color="FFFFFF")

    # Write headers
    for col_num, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Write data
    for row_num, resource in enumerate(resources, 2):
        portal_url = ""
        if tenant_id and resource.get("id"):
            portal_url = f"https://portal.azure.com/#@{tenant_id}/resource{resource['id']}"

        # Get risk level and safety info
        risk_level = resource.get("risk_level", "medium")
        info = get_resource_info(resource.get("type", ""))
        safe_to_delete = info.get("safe_to_delete", "Verify before deleting.")

        ws.cell(row=row_num, column=1, value=resource.get("name", ""))
        ws.cell(row=row_num, column=2, value=resource.get("type_display", ""))

        # Risk level cell with color
        risk_cell = ws.cell(row=row_num, column=3, value=risk_level.upper())
        risk_cell.fill = risk_fills.get(risk_level, risk_fills["medium"])
        risk_cell.font = risk_font
        risk_cell.alignment = Alignment(horizontal="center")

        ws.cell(row=row_num, column=4, value=safe_to_delete)
        ws.cell(row=row_num, column=5, value=resource.get("resource_group", ""))
        ws.cell(row=row_num, column=6, value=resource.get("location", ""))
        ws.cell(row=row_num, column=7, value=resource.get("cost_display", "$0"))
        ws.cell(row=row_num, column=8, value=resource.get("details", ""))
        ws.cell(row=row_num, column=9, value=portal_url)

    # Auto-adjust column widths
    for col_num, _ in enumerate(columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)

        for row in ws.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(filepath)
    return filepath
